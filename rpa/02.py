from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)

wb.save("sample.xlsx")
